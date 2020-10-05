try:
    import wx
    import wx.grid
    import openpyxl
    from openpyxl.utils import column_index_from_string
    from matplotlib import pyplot as plt
except ImportError:
    raise ImportError


#LOAD THE FILES NEEDED
listings = openpyxl.load_workbook('listings_dec18.xlsx')
comments = openpyxl.load_workbook('reviews_dec18.xlsx')

#THE WORKSHEETS FOR EVERY WORKBOOK FILES
listingSheet = listings['listings_dec18']
commentSheet = comments['reviews_dec18']

#GLOBAL VARIABLES
listingList = []
cleanlinessKeywords = ['clean', 'neat', 'fresh', 'hygienic', 'taintless', 'sterile', 'sanitary', 'washed', 'flawless', 'bright', 'shiny', 'sparkling']

#LISTING CLASS
class Listing:
    def __init__(self, listingID, name, hostName, street, suburb, propType, roomType, amenities, price, availability, reviewScore):
        self.id = listingID
        self.name = name
        self.hostName = hostName
        self.street = street
        self.suburb = suburb
        self.propType = propType
        self.roomType = roomType
        self.amenities = amenities
        self.price = price
        self.availability = availability
        self.reviewScore = reviewScore
        
    def __repr__(self):
        return str(self.__dict__)

#INDEX OF PROPERTIES WE'LL ADD IN THE CLASS OBJECT
listingIDCol = column_index_from_string('A') - 1
listingName = column_index_from_string('E') - 1
hostNameCol = column_index_from_string('V') - 1
streetCol = column_index_from_string('AL') - 1
suburbCol = column_index_from_string('AN') - 1
propTypeCol = column_index_from_string('AZ') - 1
roomTypeCol = column_index_from_string('BA') - 1
amenitiesCol = column_index_from_string('BG') - 1
priceCol = column_index_from_string('BI') - 1
availabilityCol = column_index_from_string('BW') - 1
reviewScoreCol = column_index_from_string('CB')

#ITERATE THROUGH THE LISTINGS FILE, CREATE AN INSTANCE OF LISTING FOR EVERY ROW
for listing in listingSheet.iter_rows(min_row=2, max_row=100, max_col=reviewScoreCol, values_only=True):
    #CREATE AN INSTANCE OF CLASS LISTING
    property = Listing(listing[listingIDCol], listing[listingName], \
                        listing[hostNameCol], listing[streetCol], \
                        listing[suburbCol], listing[propTypeCol], \
                        listing[roomTypeCol], listing[amenitiesCol], \
                        listing[priceCol], listing[availabilityCol], \
                        listing[-1])
    #ADD THE PROPERTY TO THE LIST OF LISTINGS
    listingList.append(property)
# print(listingList)

#FUNCTIONS TO RETRIEVE LISTINGS BASED ON USER INPUTS
# def showListings(listingToDisplay):
#     # class myTable(wx.Frame):
#     #     def __init__(self, parent, id, title):
#     #         wx.Frame.__init__(self, parent, id, title)
#     #         self.parent = parent
#     #         self.initialize()
            
#     #     def initialize(self):
#     #         panel= wx.Panel(self)
            
#     #         self.sizer = wx.BoxSizer(wx.VERTICAL)
            
#     #         grid = wx.grid.Grid(self, -1)
#     #         grid.SetRowLabelSize(0)
#     #         grid.SetColLabelSize(0)
#     #         grid.CreateGrid(10, len(listingToDisplay))
            
#     #         #sets the size in pixels
#     #         # grid.SetColSize(0, 100)
#     #         # grid.SetRowSize(row, height)
            
#     #         #automatically sizes the rows and columns based on the content
#     #         grid.AutoSizeColumns(setAsMin=True)
#     #         grid.AutoSizeRows(setAsMin=True)
            
#     #         #rows and columns index start from 0
#     #         # grid.SetCellValue(3, 3, 'green on grey')
            
#     #         #adding values from ext file to the grid
#     #         for listing in listingToDisplay:
#     #             rowTitle = list(listing.keys())
#     #             rowNum = 0
#     #             # colNum = 0
                
#     #             for col in range(0, 11):
#     #                 grid.SetCellValue(rowNum, col, rowTitle[col])
#     #             rowNum += 1
            
#     #         self.sizer.Add(grid)
#     #         panel.SetSizerAndFit(self.sizer)
            
#     # if __name__ == '__main__':
#     #     app = wx.App()
#     #     frame = myTable(None, wx.ID_ANY, title='Listings List')
#     #     app.MainLoop()
#     return listingToDisplay
    
# cleanCommentTotal = 0
# def findKeyword(period, keyword):
#     listingToShow = []
#     numOfCleanlinessComments = 0
#     #CHECK NUMBER OF KEYWORDS PROVIDED
#     numOfKeywords = len(keyword)
#     if numOfKeywords > 1:
#         for kw in keyword:
#             #CHECK IF KEYWORD MATCHES ANY AMENITIES
#             for listing in listingList:
#                 if (kw in listing.amenities) and (listing not in listingToShow):
#                     #ADD THE LISTINGID TO THE LIST
#                     listingToShow.append(listing)
                    
#             #CHECK IF KEYWORD MATCHES ANY COMMENTS
#             for comment in commentSheet.iter_rows(min_row=2, max_row=100, min_col=6, max_col=6, values_only=True):
#                 if kw in str(comment):
#                     listingToShow.append(listing)
            
#             #CHECK IF KEYWORD MATCHES ANY KEYWORDS ABOUT CLEANLINESS
#             for comment in cleanlinessKeywords:
#                 if kw == comment:
#                     numOfCleanlinessComments += 1

#             cleanCommentTotal = numOfCleanlinessComments
#             print('cleanCommentTotal = {}'.format(cleanCommentTotal))
#         return listingToShow
    
#     else:
#         for listing in listingList:
#             if (keyword[0] in listing.amenities) and (listing not in listingToShow):
#                 listingToShow.append(listing)     
        
#         for comment in commentSheet.iter_rows(min_row=2, max_row=100, min_col=6, max_col=6, values_only=True):
#             if keyword[0] in str(comment):
#                 listingToShow.append(listing)
#     return listingToShow

# def getListings(period, suburbName='sydney', keyword=None):
#     listingToShow = []
#     if keyword != None:
#         splitKeyword = keyword.split(', ')
#         filteredListings = findKeyword(period, splitKeyword)
        
#         #CHECK IF ANY MATCH FOUND
#         if len(filteredListings > 0):
#             for listing in filteredListings:
#                 #CHECK IF FILTERED LISTINGS' SUBURB MATCHES THE PROVIDED ARGUMENT
#                 streetList = (listing.street.lower()).split(', ')
#                 if (suburbName.lower() in streetList) or (suburbName.lower() == listing.suburb.lower()):
#                     listingToShow.append(listing)
#             return listingToShow

#         else:
#             return 'No match found'
        
#     else:
#         #ITERATE THROUGH THE GLOBAL LISTINGS LIST, CHECK IF THEY MATCH THE SUBURB PROVIDED
#         for listing in listingList:
#             #IF PROVIDED SUBURBNAME MATCHES THE SUBURB/NEIGHBOURHOOD NAME
#             if listing.suburb.lower() == suburbName.lower():
#                 listingToShow.append(listing)
    
#         return showListings(listingToShow)

# print(getListings(None))

def showPriceDist(period):
    counter = 0
    propPrices = []
    # propYears = []
    while counter <= 10:
        propPrices = [prop.price for prop in listingList]
        counter += 1
    return propPrices
    
print('prices: ', showPriceDist(None))

def showPopularListings(period, suburbName='sydney'):
    counter = 0
    propRating = []
    # propYears = []
    while counter <= 10:
        propRating = [prop.reviewScore for prop in listingList if prop.suburb.lower() == suburbName.lower()]
        counter += 1
    return propRating
    
print('ratings: ', showPopularListings(None))